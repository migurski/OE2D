'''Read tabular data from a page of a source file.

Usage: source_table.py <filename> <page_number>

Page numbers are 1-based. For XLSX/XLS files, page = sheet number.
For PDF files, page = PDF page number.
'''
from __future__ import annotations

import argparse
import bisect
import collections
import csv
import functools
import os
import sys
import typing
from xml.etree import ElementTree

import openpyxl
import pdfplumber
import pydantic
import xlrd


class BBox(pydantic.BaseModel):
    '''Bounding box compatible with pdfplumber crop() and within_bbox().'''
    x0: float = pydantic.Field(description='Left edge x coordinate')
    top: float = pydantic.Field(description='Top edge y coordinate')
    x1: float = pydantic.Field(description='Right edge x coordinate')
    bottom: float = pydantic.Field(description='Bottom edge y coordinate')


class PageTable(pydantic.BaseModel):
    '''A table found on a PDF page with its bounding box and content preview.'''
    index: int = pydantic.Field(description='Zero-based index of this table on the page')
    bbox: BBox = pydantic.Field(description='Bounding box region of the table on the page')
    row_count: int = pydantic.Field(description='Number of rows in the table')
    col_count: int = pydantic.Field(description='Number of columns in the table')
    preview: list[list[str | None]] = pydantic.Field(description='First 3 rows of table content as preview')
    strategy: str = pydantic.Field(description='Extraction strategy that found this table: "lines", "lines_strict", or "text"')


@functools.lru_cache(maxsize=None)
def _open_xlsx_workbook(path: str) -> 'openpyxl.Workbook':
    '''Open and cache an XLSX workbook (non-read-only, to support merges).'''
    return openpyxl.load_workbook(path, data_only=True, read_only=False)


@functools.lru_cache(maxsize=None)
def _open_pdf(path: str) -> 'pdfplumber.PDF':
    '''Open and cache a pdfplumber PDF.'''
    return pdfplumber.open(path)


@functools.lru_cache(maxsize=None)
def _open_xlrd_workbook(path: str) -> 'xlrd.Book':
    '''Open and cache an xlrd workbook.'''
    return xlrd.open_workbook(path)


def read_xlsx_page(path: str, page: int) -> list[list[str]] | None:
    '''Read a sheet from an XLSX file, return rows as lists of strings.

    Expands merged cells so that a value spanning multiple columns is
    repeated into each spanned column.  Uses a cached workbook so that
    reading multiple sheets doesn't re-parse the entire file.
    '''
    wb: openpyxl.Workbook = _open_xlsx_workbook(path)
    num_sheets: int = len(wb.worksheets)

    if page < 1 or page > num_sheets:
        print(f'Page {page} out of range (1-{num_sheets})', file=sys.stderr)
        return None

    ws: openpyxl.worksheet.worksheet.Worksheet = wb.worksheets[page - 1]

    # Build a map of (row, col) -> value for merged cell regions,
    # so every cell in a merge gets the top-left cell's value
    merge_fill: dict[tuple[int, int], str] = {}
    for merge_range in ws.merged_cells.ranges:
        top_left_value: str = str(ws.cell(merge_range.min_row, merge_range.min_col).value or '')
        for row_idx in range(merge_range.min_row, merge_range.max_row + 1):
            for col_idx in range(merge_range.min_col, merge_range.max_col + 1):
                merge_fill[(row_idx, col_idx)] = top_left_value

    rows: list[list[str]] = []
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        cells: list[str] = []
        for col_idx, v in enumerate(row, start=1):
            if v is not None:
                cells.append(str(v))
            elif (row_idx, col_idx) in merge_fill:
                cells.append(merge_fill[(row_idx, col_idx)])
            else:
                cells.append('')
        rows.append(cells)
    return rows


def read_xls_page(path: str, page: int) -> list[list[str]] | None:
    '''Read a sheet from an XLS file, return rows as lists of strings.

    Handles both binary BIFF format (via xlrd) and XML Spreadsheet format.
    '''
    with open(path, 'rb') as f:
        head: bytes = f.read(20)

    if head.lstrip(b'\xef\xbb\xbf').startswith(b'<?xml'):
        return _read_xml_spreadsheet_page(path, page)

    return _read_xlrd_page(path, page)


def _read_xlrd_page(path: str, page: int) -> list[list[str]] | None:
    '''Read a sheet from a binary XLS file using xlrd.'''
    wb: xlrd.Book = _open_xlrd_workbook(path)
    if page < 1 or page > wb.nsheets:
        print(f'Page {page} out of range (1-{wb.nsheets})', file=sys.stderr)
        return None
    ws: xlrd.sheet.Sheet = wb.sheet_by_index(page - 1)
    rows: list[list[str]] = []
    for i in range(ws.nrows):
        rows.append([str(ws.cell_value(i, j)) if ws.cell_value(i, j) != '' else ''
                      for j in range(ws.ncols)])
    return rows


def _read_xml_spreadsheet_page(path: str, page: int) -> list[list[str]] | None:
    '''Read a sheet from an XML Spreadsheet (SpreadsheetML) file.'''
    ns: dict[str, str] = {'s': 'urn:schemas-microsoft-com:office:spreadsheet'}
    tree: ElementTree.ElementTree = ElementTree.parse(path)
    root: ElementTree.Element = tree.getroot()
    sheets: list[ElementTree.Element] = root.findall('.//s:Worksheet', ns)

    if page < 1 or page > len(sheets):
        print(f'Page {page} out of range (1-{len(sheets)})', file=sys.stderr)
        return None

    ws: ElementTree.Element = sheets[page - 1]
    rows: list[list[str]] = []
    for row_el in ws.findall('.//s:Row', ns):
        cells: list[str] = []
        col_index: int = 0
        for cell_el in row_el.findall('s:Cell', ns):
            # Handle ss:Index attribute for skipped columns
            index_attr: str | None = cell_el.attrib.get(
                '{urn:schemas-microsoft-com:office:spreadsheet}Index'
            )
            if index_attr is not None:
                target: int = int(index_attr) - 1
                while col_index < target:
                    cells.append('')
                    col_index += 1
            data_el: ElementTree.Element | None = cell_el.find('s:Data', ns)
            cells.append(data_el.text if data_el is not None and data_el.text else '')
            col_index += 1
        rows.append(cells)
    return rows


def _read_pdf_contest_titles(pdf_page: object) -> list[list[str]] | None:
    '''Read contest title text and map each column to its contest.

    Election PDFs often have one or more contest titles (e.g. "President
    and Vice President", "Proposition 6") centered above their columns.
    This function finds those titles in the header area and assigns each
    table column to the nearest title by x-midpoint.

    Returns one row per line of title text, each with contest names
    aligned to the vertical line columns, or None if no titles are found.
    '''
    chars: list[dict] = pdf_page.chars  # type: ignore[attr-defined]
    lines: list[dict] = pdf_page.lines  # type: ignore[attr-defined]

    # Find vertical line x positions to define column boundaries
    v_xs: list[int] = sorted(set(
        round(l['x0']) for l in lines if abs(l['x0'] - l['x1']) < 1
    ))
    if not v_xs or len(v_xs) < 2:
        return None

    words: list[dict] = pdf_page.extract_words(  # type: ignore[attr-defined]
        keep_blank_chars=True, y_tolerance=1, extra_attrs=['upright'],
    )
    # Title words sit between the page header (county/date) and the
    # column headers — typically in the y range 40-100. Only consider
    # upright text to avoid picking up rotated column header chars.
    title_words: list[dict] = [
        w for w in words
        if 40 < w['top'] < 100 and w.get('upright') is True
    ]
    if not title_words:
        return None

    # Group title words by y-line (within tolerance of 5)
    y_lines: list[list[dict]] = []
    for w in sorted(title_words, key=lambda w: w['top']):
        if y_lines and abs(w['top'] - y_lines[-1][0]['top']) < 5:
            y_lines[-1].append(w)
        else:
            y_lines.append([w])

    # Build one row per y-line, assigning columns to nearest title
    num_cols: int = len(v_xs) - 1
    title_rows: list[list[str]] = []
    for line_words in y_lines:
        title_mids: list[tuple[float, str]] = [
            ((w['x0'] + w['x1']) / 2, w['text'].strip())
            for w in sorted(line_words, key=lambda w: w['x0'])
        ]
        row: list[str] = []
        for ci in range(num_cols):
            col_mid: float = (v_xs[ci] + v_xs[ci + 1]) / 2
            best_title: str = min(title_mids, key=lambda tm: abs(col_mid - tm[0]))[1]
            row.append(best_title)
        title_rows.append(row)

    return title_rows if title_rows else None


def _read_pdf_vertical_headers(pdf_page: object) -> list[str] | None:
    '''Reconstruct column headers from rotated (non-upright) PDF text.

    Many election PDFs use vertical text for column headers. This function
    detects non-upright characters, groups them by which vertical-line
    column they fall in, reads each group top-to-bottom, and assembles
    the full header label for each column.

    Returns a header row aligned to the vertical line columns, or None
    if no rotated text is found.
    '''
    chars: list[dict] = pdf_page.chars  # type: ignore[attr-defined]
    lines: list[dict] = pdf_page.lines  # type: ignore[attr-defined]

    # Find vertical line x positions to define column boundaries
    v_xs: list[int] = sorted(set(
        round(l['x0']) for l in lines if abs(l['x0'] - l['x1']) < 1
    ))
    if not v_xs:
        return None

    # Find the boundary between header and data: use the lowest
    # horizontal line that's in the top half of the page
    h_lines: list[dict] = [l for l in lines if abs(l['top'] - l['bottom']) < 1]
    page_height: float = pdf_page.height  # type: ignore[attr-defined]
    header_bottom: float = max(
        (l['top'] for l in h_lines if l['top'] < page_height / 2),
        default=page_height / 4,
    )

    # Get non-upright (rotated) characters in the header area
    rotated_chars: list[dict] = [
        c for c in chars
        if c['top'] < header_bottom and not c.get('upright')
    ]
    if not rotated_chars:
        return None

    # Detect rotation direction from the transformation matrix.
    # matrix[1] (the b component) indicates:
    #   b == -1: counterclockwise 90° — text reads top-to-bottom,
    #            multi-line headers have higher x0 = first line
    #   b ==  1: clockwise 90° — text reads bottom-to-top,
    #            multi-line headers have lower x0 = first line
    sample_b: float = rotated_chars[0].get('matrix', (0, -1))[1]
    clockwise: bool = sample_b > 0
    char_reverse: bool = clockwise
    x0_reverse: bool = not clockwise

    # Group rotated chars by which vertical-line column they fall in
    col_chars: dict[int, list[dict]] = collections.defaultdict(list)
    for c in rotated_chars:
        col_idx: int = bisect.bisect_right(v_xs, c['x0']) - 1
        col_chars[col_idx].append(c)

    # Build header row with one entry per column
    num_cols: int = len(v_xs) - 1
    headers: list[str] = [''] * num_cols
    for col_idx, col_group in col_chars.items():
        if col_idx < 0 or col_idx >= num_cols:
            continue
        # Sub-group by x0 position (each x0 is one line of the header)
        sub_groups: dict[int, list[dict]] = collections.defaultdict(list)
        for c in col_group:
            sub_groups[round(c['x0'])].append(c)
        # Read each sub-group in character order, then join lines
        text_lines: list[str] = []
        for x0 in sorted(sub_groups.keys(), reverse=x0_reverse):
            line_text: str = ''.join(
                c['text'] for c in sorted(
                    sub_groups[x0], key=lambda c: c['top'], reverse=char_reverse,
                )
            ).strip()
            if line_text:
                text_lines.append(line_text)
        headers[col_idx] = ' '.join(text_lines)

    return headers


def read_pdf_page(path: str, page: int) -> list[list[str]] | None:
    '''Read a table from a PDF page, return rows as lists of strings.

    Uses text-based horizontal line detection so that each text line
    becomes its own row, rather than merging everything between the
    sparse horizontal rules into a single cell. Reconstructs column
    headers from rotated (vertical) text when present.
    '''
    pdf: pdfplumber.PDF = _open_pdf(path)
    if page < 1 or page > len(pdf.pages):
        print(f'Page {page} out of range (1-{len(pdf.pages)})', file=sys.stderr)
        return None

    pdf_page: pdfplumber.page.Page = pdf.pages[page - 1]

    # Check whether the page has vertical lines for column boundaries
    page_lines: list[dict] = pdf_page.lines
    has_vertical_lines: bool = any(
        abs(l['x0'] - l['x1']) < 1 for l in page_lines
    )

    table_settings: dict[str, object] = {
        'horizontal_strategy': 'text',
        'snap_y_tolerance': 4,
        'snap_x_tolerance': 4,
        'join_y_tolerance': 4,
        'join_x_tolerance': 4,
    }

    if has_vertical_lines:
        table_settings['vertical_strategy'] = 'lines'
    else:
        # No vertical lines — derive column boundaries from horizontal
        # line segment endpoints. Use the horizontal line y-position with
        # the most segments (the most detailed column breakdown), then
        # merge nearby x-values (within 5px) to collapse inter-group gaps.
        h_lines: list[dict] = [
            l for l in page_lines if abs(l['top'] - l['bottom']) < 1
        ]
        y_groups: dict[float, list[dict]] = collections.defaultdict(list)
        for l in h_lines:
            y_groups[round(l['top'], 1)].append(l)
        multi_seg_ys: list[float] = [
            y for y, segs in y_groups.items() if len(segs) > 1
        ]
        if not multi_seg_ys:
            print(f'No column structure found on page {page}', file=sys.stderr)
            return None
        best_y: float = max(multi_seg_ys, key=lambda y: len(y_groups[y]))
        best_segs: list[dict] = y_groups[best_y]
        x_points: list[float] = sorted(set(
            round(l['x0'], 1) for l in best_segs
        ) | set(
            round(l['x1'], 1) for l in best_segs
        ))
        merged_xs: list[float] = []
        for x in x_points:
            if merged_xs and x - merged_xs[-1] < 5:
                merged_xs[-1] = (merged_xs[-1] + x) / 2
            else:
                merged_xs.append(x)
        table_settings['vertical_strategy'] = 'explicit'
        table_settings['explicit_vertical_lines'] = merged_xs

    tables: list[list[list[str | None]]] = pdf_page.extract_tables(table_settings)
    if not tables:
        print(f'No tables found on page {page}', file=sys.stderr)
        return None

    # Find the most common column width — that's the data table width.
    # Concatenate all tables with that width to capture stacked contests.
    width_counts: dict[int, int] = collections.defaultdict(int)
    for t in tables:
        if t:
            width_counts[len(t[0])] += len(t)
    data_width: int = max(width_counts, key=width_counts.get)
    table: list[list[str | None]] = []
    for t in tables:
        if t and len(t[0]) == data_width:
            table.extend(t)

    # Reconstruct headers from vertical text if present
    vertical_headers: list[str] | None = _read_pdf_vertical_headers(pdf_page)
    contest_titles: list[list[str]] | None = _read_pdf_contest_titles(pdf_page)

    rows: list[list[str]] = []
    for row in table:
        cells: list[str] = [str(v) if v is not None else '' for v in row]
        # Skip blank rows
        if any(c.strip() for c in cells):
            rows.append(cells)

    # Replace garbled header rows with reconstructed vertical headers
    if vertical_headers and rows:
        # Find where data rows start: first row where both the first and
        # second cells are non-empty and contain no newlines (garbled header
        # rows have empty cells, newline fragments, or span columns)
        data_start: int = 0
        for i, row in enumerate(rows):
            c0: str = row[0].strip() if row else ''
            c1: str = row[1].strip() if len(row) > 1 else ''
            if c0 and '\n' not in c0 and c1 and '\n' not in c1:
                data_start = i
                break

        # Pad headers to match data row width
        data_width: int = len(rows[data_start]) if data_start < len(rows) else len(vertical_headers)
        while len(vertical_headers) < data_width:
            vertical_headers.append('')

        header_rows: list[list[str]] = []

        # Add contest title rows if present
        if contest_titles:
            for title_row in contest_titles:
                while len(title_row) < data_width:
                    title_row.append('')
                header_rows.append(title_row[:data_width])

        header_rows.append(vertical_headers[:data_width])
        rows = header_rows + rows[data_start:]

    return rows


@functools.lru_cache(maxsize=None)
def page_count(path: str) -> int:
    '''Return the number of pages (PDF) or sheets (XLSX/XLS) in a file.'''
    ext: str = os.path.splitext(path)[1].lower()

    if ext == '.xlsx':
        return len(_open_xlsx_workbook(path).worksheets)

    if ext == '.xls':
        with open(path, 'rb') as f:
            head: bytes = f.read(20)
        if head.lstrip(b'\xef\xbb\xbf').startswith(b'<?xml'):
            ns: dict[str, str] = {'s': 'urn:schemas-microsoft-com:office:spreadsheet'}
            tree: ElementTree.ElementTree = ElementTree.parse(path)
            return len(tree.getroot().findall('.//s:Worksheet', ns))
        return _open_xlrd_workbook(path).nsheets

    if ext == '.pdf':
        return len(_open_pdf(path).pages)

    return 0


@functools.lru_cache(maxsize=None)
def page_table(path: str, page: int) -> list[list[str]] | None:
    '''Read tabular data from a page of a source file.

    Routes to the appropriate reader based on file extension.
    Returns rows as lists of strings, or None on failure.
    '''
    ext: str = os.path.splitext(path)[1].lower()

    if ext == '.xlsx':
        return read_xlsx_page(path, page)
    elif ext == '.xls':
        return read_xls_page(path, page)
    elif ext == '.pdf':
        return read_pdf_page(path, page)
    else:
        print(f'Unsupported file type: {ext}', file=sys.stderr)
        return None


def page_tables(
    path: str,
    page: int,
    strategy: typing.Literal['lines', 'lines_strict', 'text'] = 'lines',
) -> list[PageTable] | None:
    '''Find tables on a PDF page using the given extraction strategy.

    Returns a list of PageTable instances with bounding boxes and content
    previews, or None if the file is not a PDF, the page is out of range,
    or no tables are found.

    The strategy parameter sets both vertical and horizontal extraction:
    "lines" uses ruled lines to find table boundaries, "lines_strict"
    uses only explicit line intersections, and "text" uses text alignment
    to infer column structure.
    '''
    ext: str = os.path.splitext(path)[1].lower()
    if ext != '.pdf':
        return None

    pdf: pdfplumber.PDF = _open_pdf(path)
    if page < 1 or page > len(pdf.pages):
        return None

    pdf_page: pdfplumber.page.Page = pdf.pages[page - 1]

    table_settings: dict[str, str] = {
        'vertical_strategy': strategy,
        'horizontal_strategy': strategy,
    }

    tables = pdf_page.find_tables(table_settings)
    if not tables:
        return None

    result: list[PageTable] = []
    for i, table in enumerate(tables):
        rows = table.extract()
        result.append(PageTable(
            index=i,
            bbox=BBox(
                x0=round(table.bbox[0], 1),
                top=round(table.bbox[1], 1),
                x1=round(table.bbox[2], 1),
                bottom=round(table.bbox[3], 1),
            ),
            row_count=len(rows),
            col_count=len(rows[0]) if rows else 0,
            preview=rows[:3],
            strategy=strategy,
        ))

    return result


def page_words(path: str, page: int) -> list[dict] | None:
    '''Extract words with positions from a PDF page.

    Returns a list of word dicts with keys: text, x0, x1, top, bottom,
    upright. Useful for scanning page content and finding contest names
    or candidate names without committing to a table extraction strategy.

    Returns None if the file is not a PDF or the page is out of range.
    '''
    ext: str = os.path.splitext(path)[1].lower()
    if ext != '.pdf':
        return None

    pdf: pdfplumber.PDF = _open_pdf(path)
    if page < 1 or page > len(pdf.pages):
        return None

    pdf_page: pdfplumber.page.Page = pdf.pages[page - 1]
    words = pdf_page.extract_words(keep_blank_chars=True, extra_attrs=['upright'])

    return [
        {
            'text': w['text'],
            'x0': round(w['x0'], 1),
            'x1': round(w['x1'], 1),
            'top': round(w['top'], 1),
            'bottom': round(w['bottom'], 1),
            'upright': w['upright'],
        }
        for w in words
    ]


def main() -> None:
    parser: argparse.ArgumentParser = argparse.ArgumentParser(
        description='Read tabular data from a page of a source file.',
    )
    parser.add_argument('filename', help='Path to the source file')
    parser.add_argument('page', type=int, help='Page number (1-based)')
    args: argparse.Namespace = parser.parse_args()

    rows: list[list[str]] | None = page_table(args.filename, args.page)

    if rows is None:
        sys.exit(1)

    try:
        writer: csv.writer = csv.writer(sys.stdout)
        for row in rows:
            writer.writerow(row)
        sys.stdout.flush()
    except BrokenPipeError:
        pass


if __name__ == '__main__':
    main()
