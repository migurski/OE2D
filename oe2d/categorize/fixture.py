'''Trim a large election source into a small categorization fixture.

Usage:
  oe2d-make-fixture <source-url-or-path> [--out DIR] [--pages N] [--sheets M] [--rows R]
  oe2d-make-fixture --manifest oe2d/categorize/labels/seed_sources.tsv [--out DIR] [--limit N]

Election sources routinely run to hundreds of pages and tens of thousands of
rows. For categorization we only need enough to see the shape — the container
format, the header/contest structure, and a few data rows — like the excerpted
CA files in ./fixtures. This tool downloads a source (or reads a local one),
keeps a handful of representative pages/sheets/rows while preserving the
container format, and writes a small fixture next to a JSON summary.
'''
from __future__ import annotations

import argparse
import csv
import json
import os
import re
import shutil
import sys
import tempfile
import urllib.parse
import urllib.request
import xml.etree.ElementTree as ET
import zipfile

import openpyxl
import pypdf
import xlrd
import xlwt

import source_table
from .. import categorize

# SpreadsheetML namespace for XML-format .xls files.
_SS_NS = 'urn:schemas-microsoft-com:office:spreadsheet'

# xlwt hard limits for the legacy BIFF format.
_XLS_MAX_COLS = 256
_XLS_MAX_SHEETNAME = 31

# Columns kept per sheet when trimming — enough to see candidate orientation
# without carrying a contest's full 200+ column width.
_TRIM_MAX_COLS = 40

# Scanned pages are large images; drop pages until a PDF fixture fits this.
_PDF_MAX_BYTES = 3_000_000

# Copy spreadsheets whole below this size so they preview natively; re-serializing
# a workbook strips sharedStrings/activeTab and can leave Quick Look blank, and
# these one-contest-per-sheet workbooks often lead with a table-of-contents sheet.
_SPREADSHEET_MAX_BYTES = 2_000_000
_SPREADSHEET_CONTAINERS = ('xlsx', 'xls_binary', 'xls_xml')


def slugify(name: str) -> str:
    '''Turn a source basename into a safe fixture stem.'''
    stem: str = os.path.splitext(name)[0].lower()
    stem = re.sub(r'[^a-z0-9]+', '-', stem).strip('-')
    return stem[:80] or 'fixture'


def fetch(source: str, work_dir: str) -> str:
    '''Return a local path for a source given as a URL or a path.'''
    if source.startswith(('http://', 'https://')):
        name: str = os.path.basename(urllib.parse.urlparse(source).path)
        local: str = os.path.join(work_dir, urllib.parse.unquote(name))
        with urllib.request.urlopen(source) as response, open(local, 'wb') as out:
            shutil.copyfileobj(response, out)
        return local
    return source


def _pdf_page_indices(total: int, pages: int) -> list[int]:
    '''Pick the first pages-1 pages plus one from the middle (deduped).

    Early pages carry statistics/first contests; a middle page samples a later
    contest, mirroring how the CA fixtures grabbed non-adjacent pages.
    '''
    indices: list[int] = list(range(min(pages - 1, total)))
    if total > pages:
        indices.append(total // 2)
    return sorted(set(indices))[:pages] if pages > 0 else []


def _excerpt_pdf(path: str, out_path: str, pages: int) -> None:
    '''Keep a few pages, shedding more if a scanned fixture stays too large.'''
    reader: pypdf.PdfReader = pypdf.PdfReader(path)
    total: int = len(reader.pages)
    kept: int = pages
    while True:
        writer: pypdf.PdfWriter = pypdf.PdfWriter()
        for i in _pdf_page_indices(total, kept):
            writer.add_page(reader.pages[i])
        with open(out_path, 'wb') as out:
            writer.write(out)
        if kept <= 1 or os.path.getsize(out_path) <= _PDF_MAX_BYTES:
            return
        kept -= 1


def _excerpt_xlsx(path: str, out_path: str, sheets: int, rows: int) -> None:
    '''Rebuild the first sheets/rows into a fresh workbook.

    Deleting trailing rows leaves the saved sheet dimension unshrunk, so the
    trimmed rows read back anyway. Copying values into a new workbook gives a
    genuinely small sheet.
    '''
    source: openpyxl.Workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out_book: openpyxl.Workbook = openpyxl.Workbook()
    out_book.remove(out_book.active)
    for worksheet in source.worksheets[:sheets]:
        out_sheet = out_book.create_sheet(worksheet.title[:_XLS_MAX_SHEETNAME])
        for _, row in zip(range(rows), worksheet.iter_rows(values_only=True)):
            out_sheet.append(list(row)[:_TRIM_MAX_COLS])
    out_book.save(out_path)
    source.close()


def _excerpt_lines(path: str, out_path: str, rows: int) -> None:
    with open(path, encoding='utf-8', errors='replace') as src, \
            open(out_path, 'w', encoding='utf-8') as out:
        for _, line in zip(range(rows), src):
            out.write(line)


def _excerpt_xls_binary(path: str, out_path: str, sheets: int, rows: int) -> None:
    '''Trim a binary BIFF .xls by rewriting a subset with xlwt.'''
    book: xlrd.Book = xlrd.open_workbook(path)
    out_book: xlwt.Workbook = xlwt.Workbook()
    used_names: set[str] = set()
    for sheet_index in range(min(sheets, book.nsheets)):
        sheet: xlrd.sheet.Sheet = book.sheet_by_index(sheet_index)
        sheet_name: str = (sheet.name or f'Sheet{sheet_index + 1}')[:_XLS_MAX_SHEETNAME]
        while sheet_name in used_names:
            sheet_name = f'{sheet_name[:_XLS_MAX_SHEETNAME - 2]}_{sheet_index}'
        used_names.add(sheet_name)
        out_sheet = out_book.add_sheet(sheet_name, cell_overwrite_ok=True)
        for row_index in range(min(rows, sheet.nrows)):
            for col_index in range(min(_TRIM_MAX_COLS, _XLS_MAX_COLS, sheet.ncols)):
                out_sheet.write(row_index, col_index, sheet.cell_value(row_index, col_index))
    out_book.save(out_path)


def _excerpt_xls_xml(path: str, out_path: str, sheets: int, rows: int) -> None:
    '''Trim an XML SpreadsheetML .xls by pruning worksheets and rows.'''
    tree: ET.ElementTree = ET.parse(path)
    root: ET.Element = tree.getroot()
    worksheets: list[ET.Element] = root.findall(f'{{{_SS_NS}}}Worksheet')
    for extra in worksheets[sheets:]:
        root.remove(extra)
    for worksheet in worksheets[:sheets]:
        table: ET.Element | None = worksheet.find(f'{{{_SS_NS}}}Table')
        if table is None:
            continue
        kept: int = 0
        for row in list(table.findall(f'{{{_SS_NS}}}Row')):
            kept += 1
            if kept > rows:
                table.remove(row)
                continue
            for extra_cell in row.findall(f'{{{_SS_NS}}}Cell')[_TRIM_MAX_COLS:]:
                row.remove(extra_cell)
        expanded: str = f'{{{_SS_NS}}}ExpandedRowCount'
        if expanded in table.attrib:
            table.attrib[expanded] = str(min(rows, kept))
    ET.register_namespace('ss', _SS_NS)
    tree.write(out_path, xml_declaration=True, encoding='utf-8')


def _excerpt_zip(path: str, out_path: str, pages: int, sheets: int, rows: int,
                 members: int) -> None:
    '''Keep the first members of a zip bundle, each trimmed in turn.'''
    with zipfile.ZipFile(path) as archive, \
            zipfile.ZipFile(out_path, 'w', zipfile.ZIP_DEFLATED) as out_archive, \
            tempfile.TemporaryDirectory() as work_dir:
        kept: int = 0
        for info in archive.infolist():
            if info.is_dir():
                continue
            if kept >= members:
                break
            kept += 1
            member: str = archive.extract(info, work_dir)
            trimmed: str = os.path.join(work_dir, 'trimmed-' + os.path.basename(member))
            try:
                _trim_local(member, trimmed, categorize.detect_container(member),
                            pages, sheets, rows, members)
            except Exception:
                trimmed = member
            out_archive.write(trimmed, arcname=info.filename)


def _trim_local(local: str, out_path: str, container: str,
                pages: int, sheets: int, rows: int, members: int) -> None:
    '''Dispatch to the right excerpt strategy for a local file's container.'''
    if container in ('vector_pdf', 'scanned_pdf'):
        _excerpt_pdf(local, out_path, pages)
    elif container == 'xlsx':
        _excerpt_xlsx(local, out_path, sheets, rows)
    elif container == 'xls_binary':
        _excerpt_xls_binary(local, out_path, sheets, rows)
    elif container == 'xls_xml':
        _excerpt_xls_xml(local, out_path, sheets, rows)
    elif container == 'zip':
        _excerpt_zip(local, out_path, pages, sheets, rows, members)
    elif container in ('csv', 'txt'):
        _excerpt_lines(local, out_path, rows)
    else:  # docx, unknown — small archives kept whole
        shutil.copyfile(local, out_path)


def excerpt(source: str, out_dir: str, name: str | None = None,
            pages: int = 4, sheets: int = 8, rows: int = 60, members: int = 3,
            spreadsheet_max_bytes: int = _SPREADSHEET_MAX_BYTES) -> str:
    '''Write a fixture for a source, returning the output path.

    Small spreadsheets are copied whole so they preview natively; everything
    else (and oversized spreadsheets) is trimmed per container.
    '''
    os.makedirs(out_dir, exist_ok=True)
    with tempfile.TemporaryDirectory() as work_dir:
        local: str = fetch(source, work_dir)
        container: str = categorize.detect_container(local)
        ext: str = os.path.splitext(local)[1].lower()
        stem: str = name or slugify(os.path.basename(local))
        out_path: str = os.path.join(out_dir, stem + ext)
        if (container in _SPREADSHEET_CONTAINERS
                and os.path.getsize(local) <= spreadsheet_max_bytes):
            shutil.copyfile(local, out_path)
        else:
            _trim_local(local, out_path, container, pages, sheets, rows, members)
        return out_path


def _summarize(out_path: str) -> dict:
    '''Report a produced fixture's container, page count, and size.

    Uses the deterministic detectors only — no RLM/model call, since this just
    describes what was written.
    '''
    source_table.page_count.cache_clear()
    container: str = categorize.detect_container(out_path)
    return {
        'file_name': os.path.basename(out_path),
        'container': container,
        'page_count': categorize.count_pages(out_path, container),
        'bytes': os.path.getsize(out_path),
    }


def _run_manifest(manifest: str, out_dir: str, limit: int | None,
                  pages: int, sheets: int, rows: int, members: int,
                  spreadsheet_max_bytes: int) -> None:
    with open(manifest, encoding='utf-8') as handle:
        records: list[dict] = list(csv.DictReader(handle, delimiter='\t'))
    if limit:
        records = records[:limit]
    for record in records:
        source: str = record['url']
        if not source.startswith('http'):
            print(f'skip (local): {record["file"]}', file=sys.stderr)
            continue
        try:
            out_path: str = excerpt(source, out_dir, pages=pages, sheets=sheets,
                                    rows=rows, members=members,
                                    spreadsheet_max_bytes=spreadsheet_max_bytes)
            summary: dict = _summarize(out_path)
            print(f'{summary["bytes"]:>9d}  {summary["container"]:11s}  {os.path.basename(out_path)}')
        except Exception as err:
            print(f'FAIL {record["file"]}: {err}', file=sys.stderr)


def main() -> None:
    parser: argparse.ArgumentParser = argparse.ArgumentParser(
        description='Trim an election source into a small categorization fixture.',
    )
    parser.add_argument('source', nargs='?', help='Source URL or path')
    parser.add_argument('--manifest', help='TSV of seed sources to batch-process')
    parser.add_argument('--out', default='oe2d/tests/categorize/fixtures', help='Output directory')
    parser.add_argument('--name', help='Fixture stem (single-source mode)')
    parser.add_argument('--pages', type=int, default=4, help='PDF pages to keep')
    parser.add_argument('--sheets', type=int, default=8,
                        help='Workbook sheets to keep (trim path); early sheets are often a TOC')
    parser.add_argument('--rows', type=int, default=60, help='Rows per sheet / text lines')
    parser.add_argument('--zip-members', type=int, default=3, help='Zip members to keep')
    parser.add_argument('--sheet-max-bytes', type=int, default=_SPREADSHEET_MAX_BYTES,
                        help='Copy spreadsheets whole at or below this size')
    parser.add_argument('--limit', type=int, help='Manifest rows to process')
    args: argparse.Namespace = parser.parse_args()

    if args.manifest:
        _run_manifest(args.manifest, args.out, args.limit, args.pages, args.sheets,
                      args.rows, args.zip_members, args.sheet_max_bytes)
        return

    if not args.source:
        parser.error('provide a source, or --manifest')

    out_path: str = excerpt(
        args.source, args.out, name=args.name,
        pages=args.pages, sheets=args.sheets, rows=args.rows, members=args.zip_members,
        spreadsheet_max_bytes=args.sheet_max_bytes,
    )
    print(json.dumps(_summarize(out_path), indent=2))


if __name__ == '__main__':
    main()
