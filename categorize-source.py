'''Read tabular data from a page of a source file.

Usage: categorize-source.py <filename> <page_number>

Page numbers are 1-based. For XLSX/XLS files, page = sheet number.
For PDF files, page = PDF page number.
'''
from __future__ import annotations

import argparse
import bisect
import csv
import io
import os
import sys
import xml.etree.ElementTree as ET
from collections import defaultdict


def read_xlsx_page(path: str, page: int) -> list[list[str]] | None:
    '''Read a sheet from an XLSX file, return rows as lists of strings.

    Expands merged cells so that a value spanning multiple columns is
    repeated into each spanned column.
    '''
    import openpyxl

    wb: openpyxl.Workbook = openpyxl.load_workbook(path, data_only=True)
    if page < 1 or page > len(wb.sheetnames):
        print(f'Page {page} out of range (1-{len(wb.sheetnames)})', file=sys.stderr)
        return None
    ws: openpyxl.worksheet.worksheet.Worksheet = wb.worksheets[page - 1]

    # Build a map of (row, col) -> value for merged cell regions,
    # so every cell in a merge gets the top-left cell's value
    merge_fill: dict[tuple[int, int], str] = {}
    for merge_range in ws.merged_cells.ranges:
        top_left_value: str = str(ws.cell(merge_range.min_row, merge_range.min_col).value or '')
        for row_idx in range(merge_range.min_row, merge_range.max_row + 1):
            for col_idx in range(merge_range.min_col, merge_range.max_col + 1):
                merge_fill[(row_idx, col_idx)] = top_left_value

    rows: list[list[str]] = []
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        cells: list[str] = []
        for col_idx, v in enumerate(row, start=1):
            if v is not None:
                cells.append(str(v))
            elif (row_idx, col_idx) in merge_fill:
                cells.append(merge_fill[(row_idx, col_idx)])
            else:
                cells.append('')
        rows.append(cells)
    wb.close()
    return rows


def read_xls_page(path: str, page: int) -> list[list[str]] | None:
    '''Read a sheet from an XLS file, return rows as lists of strings.

    Handles both binary BIFF format (via xlrd) and XML Spreadsheet format.
    '''
    with open(path, 'rb') as f:
        head: bytes = f.read(20)

    if head.lstrip(b'\xef\xbb\xbf').startswith(b'<?xml'):
        return _read_xml_spreadsheet_page(path, page)

    return _read_xlrd_page(path, page)


def _read_xlrd_page(path: str, page: int) -> list[list[str]] | None:
    '''Read a sheet from a binary XLS file using xlrd.'''
    import xlrd

    wb: xlrd.Book = xlrd.open_workbook(path)
    if page < 1 or page > wb.nsheets:
        print(f'Page {page} out of range (1-{wb.nsheets})', file=sys.stderr)
        return None
    ws: xlrd.sheet.Sheet = wb.sheet_by_index(page - 1)
    rows: list[list[str]] = []
    for i in range(ws.nrows):
        rows.append([str(ws.cell_value(i, j)) if ws.cell_value(i, j) != '' else ''
                      for j in range(ws.ncols)])
    return rows


def _read_xml_spreadsheet_page(path: str, page: int) -> list[list[str]] | None:
    '''Read a sheet from an XML Spreadsheet (SpreadsheetML) file.'''
    ns: dict[str, str] = {'s': 'urn:schemas-microsoft-com:office:spreadsheet'}
    tree: ET.ElementTree = ET.parse(path)
    root: ET.Element = tree.getroot()
    sheets: list[ET.Element] = root.findall('.//s:Worksheet', ns)

    if page < 1 or page > len(sheets):
        print(f'Page {page} out of range (1-{len(sheets)})', file=sys.stderr)
        return None

    ws: ET.Element = sheets[page - 1]
    rows: list[list[str]] = []
    for row_el in ws.findall('.//s:Row', ns):
        cells: list[str] = []
        col_index: int = 0
        for cell_el in row_el.findall('s:Cell', ns):
            # Handle ss:Index attribute for skipped columns
            index_attr: str | None = cell_el.attrib.get(
                '{urn:schemas-microsoft-com:office:spreadsheet}Index'
            )
            if index_attr is not None:
                target: int = int(index_attr) - 1
                while col_index < target:
                    cells.append('')
                    col_index += 1
            data_el: ET.Element | None = cell_el.find('s:Data', ns)
            cells.append(data_el.text if data_el is not None and data_el.text else '')
            col_index += 1
        rows.append(cells)
    return rows


def _read_pdf_contest_titles(pdf_page: object) -> list[str] | None:
    '''Read contest title text and map each column to its contest.

    Election PDFs often have one or more contest titles (e.g. "President
    and Vice President", "Proposition 6") centered above their columns.
    This function finds those titles in the header area and assigns each
    table column to the nearest title by x-midpoint.

    Returns a row of contest names aligned to the vertical line columns,
    or None if no titles are found.
    '''
    import pdfplumber

    chars: list[dict] = pdf_page.chars  # type: ignore[attr-defined]
    lines: list[dict] = pdf_page.lines  # type: ignore[attr-defined]

    # Find vertical line x positions to define column boundaries
    v_xs: list[int] = sorted(set(
        round(l['x0']) for l in lines if abs(l['x0'] - l['x1']) < 1
    ))
    if not v_xs or len(v_xs) < 2:
        return None

    words: list[dict] = pdf_page.extract_words(  # type: ignore[attr-defined]
        keep_blank_chars=True, y_tolerance=1,
    )
    # Title words sit between the page header (county/date) and the
    # column headers — typically in the y range 55-100
    title_words: list[dict] = [w for w in words if 55 < w['top'] < 100]
    if not title_words:
        return None

    # Filter to title-like words: the topmost group of words in that range
    # (contest names appear at the same y, "VOTE FOR 1" and "Precincts
    # Reporting" appear on lower lines)
    min_top: float = min(w['top'] for w in title_words)
    titles: list[dict] = sorted(
        [w for w in title_words if w['top'] < min_top + 5],
        key=lambda w: w['x0'],
    )
    if not titles:
        return None

    # Title midpoints
    title_mids: list[tuple[float, str]] = [
        ((t['x0'] + t['x1']) / 2, t['text'].strip()) for t in titles
    ]

    # Assign each column to the nearest title by x-midpoint
    num_cols: int = len(v_xs) - 1
    contest_row: list[str] = []
    for ci in range(num_cols):
        col_mid: float = (v_xs[ci] + v_xs[ci + 1]) / 2
        best_title: str = min(title_mids, key=lambda tm: abs(col_mid - tm[0]))[1]
        contest_row.append(best_title)

    return contest_row


def _read_pdf_vertical_headers(pdf_page: object) -> list[str] | None:
    '''Reconstruct column headers from rotated (non-upright) PDF text.

    Many election PDFs use vertical text for column headers. This function
    detects non-upright characters, groups them by which vertical-line
    column they fall in, reads each group top-to-bottom, and assembles
    the full header label for each column.

    Returns a header row aligned to the vertical line columns, or None
    if no rotated text is found.
    '''
    chars: list[dict] = pdf_page.chars  # type: ignore[attr-defined]
    lines: list[dict] = pdf_page.lines  # type: ignore[attr-defined]

    # Find vertical line x positions to define column boundaries
    v_xs: list[int] = sorted(set(
        round(l['x0']) for l in lines if abs(l['x0'] - l['x1']) < 1
    ))
    if not v_xs:
        return None

    # Find the boundary between header and data: use the lowest
    # horizontal line that's in the top half of the page
    h_lines: list[dict] = [l for l in lines if abs(l['top'] - l['bottom']) < 1]
    page_height: float = pdf_page.height  # type: ignore[attr-defined]
    header_bottom: float = max(
        (l['top'] for l in h_lines if l['top'] < page_height / 2),
        default=page_height / 4,
    )

    # Get non-upright (rotated) characters in the header area
    rotated_chars: list[dict] = [
        c for c in chars
        if c['top'] < header_bottom and not c.get('upright')
    ]
    if not rotated_chars:
        return None

    # Group rotated chars by which vertical-line column they fall in
    col_chars: dict[int, list[dict]] = defaultdict(list)
    for c in rotated_chars:
        col_idx: int = bisect.bisect_right(v_xs, c['x0']) - 1
        col_chars[col_idx].append(c)

    # Build header row with one entry per column
    num_cols: int = len(v_xs) - 1
    headers: list[str] = [''] * num_cols
    for col_idx, col_group in col_chars.items():
        if col_idx < 0 or col_idx >= num_cols:
            continue
        # Sub-group by x0 position (each x0 is one line of the header)
        sub_groups: dict[int, list[dict]] = defaultdict(list)
        for c in col_group:
            sub_groups[round(c['x0'])].append(c)
        # Read each sub-group top-to-bottom, then join in reverse x0 order
        # (rightmost x0 = topmost line of the rotated header)
        text_lines: list[str] = []
        for x0 in sorted(sub_groups.keys(), reverse=True):
            line_text: str = ''.join(
                c['text'] for c in sorted(sub_groups[x0], key=lambda c: c['top'])
            ).strip()
            if line_text:
                text_lines.append(line_text)
        headers[col_idx] = ' '.join(text_lines)

    return headers


def read_pdf_page(path: str, page: int) -> list[list[str]] | None:
    '''Read a table from a PDF page, return rows as lists of strings.

    Uses text-based horizontal line detection so that each text line
    becomes its own row, rather than merging everything between the
    sparse horizontal rules into a single cell. Reconstructs column
    headers from rotated (vertical) text when present.
    '''
    import pdfplumber

    pdf: pdfplumber.PDF = pdfplumber.open(path)
    if page < 1 or page > len(pdf.pages):
        print(f'Page {page} out of range (1-{len(pdf.pages)})', file=sys.stderr)
        return None

    pdf_page: pdfplumber.page.Page = pdf.pages[page - 1]
    table_settings: dict[str, object] = {
        'vertical_strategy': 'lines',
        'horizontal_strategy': 'text',
        'snap_y_tolerance': 4,
        'snap_x_tolerance': 4,
        'join_y_tolerance': 4,
        'join_x_tolerance': 4,
    }
    tables: list[list[list[str | None]]] = pdf_page.extract_tables(table_settings)
    if not tables:
        print(f'No tables found on page {page}', file=sys.stderr)
        return None

    # Use the largest table on the page
    table: list[list[str | None]] = max(tables, key=len)

    # Reconstruct headers from vertical text if present
    vertical_headers: list[str] | None = _read_pdf_vertical_headers(pdf_page)
    contest_titles: list[str] | None = _read_pdf_contest_titles(pdf_page)

    rows: list[list[str]] = []
    for row in table:
        cells: list[str] = [str(v) if v is not None else '' for v in row]
        # Skip blank rows
        if any(c.strip() for c in cells):
            rows.append(cells)

    # Replace garbled header rows with reconstructed vertical headers
    if vertical_headers and rows:
        # Find where data rows start: first row where both the first and
        # second cells are non-empty and contain no newlines (garbled header
        # rows have empty cells, newline fragments, or span columns)
        data_start: int = 0
        for i, row in enumerate(rows):
            c0: str = row[0].strip() if row else ''
            c1: str = row[1].strip() if len(row) > 1 else ''
            if c0 and '\n' not in c0 and c1 and '\n' not in c1:
                data_start = i
                break

        # Pad headers to match data row width
        data_width: int = len(rows[data_start]) if data_start < len(rows) else len(vertical_headers)
        while len(vertical_headers) < data_width:
            vertical_headers.append('')

        header_rows: list[list[str]] = []

        # Add contest title row if present
        if contest_titles:
            while len(contest_titles) < data_width:
                contest_titles.append('')
            header_rows.append(contest_titles[:data_width])

        header_rows.append(vertical_headers[:data_width])
        rows = header_rows + rows[data_start:]

    pdf.close()
    return rows


def page_table(path: str, page: int) -> list[list[str]] | None:
    '''Read tabular data from a page of a source file.

    Routes to the appropriate reader based on file extension.
    Returns rows as lists of strings, or None on failure.
    '''
    ext: str = os.path.splitext(path)[1].lower()

    if ext == '.xlsx':
        return read_xlsx_page(path, page)
    elif ext == '.xls':
        return read_xls_page(path, page)
    elif ext == '.pdf':
        return read_pdf_page(path, page)
    else:
        print(f'Unsupported file type: {ext}', file=sys.stderr)
        return None


def main() -> None:
    parser: argparse.ArgumentParser = argparse.ArgumentParser(
        description='Read tabular data from a page of a source file.',
    )
    parser.add_argument('filename', help='Path to the source file')
    parser.add_argument('page', type=int, help='Page number (1-based)')
    args: argparse.Namespace = parser.parse_args()

    rows: list[list[str]] | None = page_table(args.filename, args.page)

    if rows is None:
        sys.exit(1)

    try:
        writer: csv.writer = csv.writer(sys.stdout)
        for row in rows:
            writer.writerow(row)
        sys.stdout.flush()
    except BrokenPipeError:
        pass


if __name__ == '__main__':
    main()
